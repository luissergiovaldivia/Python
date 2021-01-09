import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

import os

print ("Actual directorio" , os.getcwd())

FILE_PATH='ejemplo1.xlsx'
SHEET='Hoja1'

#os.chdir("c:\\Users\\Sofia\\Documents")

#print ("Cambio de directorio" , os.getcwd())


#wb=load_workbook(FILE_PATH, read_only=True)
wb=load_workbook(FILE_PATH, read_only=0)
print(wb)

print(wb.get_sheet_names())

#print(wb.get_sheet_name())
#wb.get_sheet_names()

#sheet = wb.get_sheet_by_name('Hoja3')
#sheet.title

sheet = wb[SHEET]
#type(sheet)
print(sheet)

sheet['A1'].value
print (sheet['A1'].value)

sheet['B1'].value
print (sheet['B1'].value)


sheet['C1'].value
print (sheet['C1'].value)


sheet['C1'] = 42
sheet['C1'].value

sheet.title