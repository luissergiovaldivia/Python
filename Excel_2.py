from openpyxl import load_workbook

FILE_PATH = 'ejemplo1.xlsx'

SHEET = 'Hoja1'

workbook = load_workbook(FILE_PATH, read_only=True)

sheet = workbook[SHEET]

for Fecha, Producto, Cantidad in sheet.iter_rows():
	print(Fecha.value)
	print(Producto.value)
	print(Cantidad.value)
	#print(row[3].value)